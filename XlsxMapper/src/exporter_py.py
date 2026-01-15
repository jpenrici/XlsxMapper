# -*- coding: utf-8 -*-

import hashlib

from pathlib import Path
from typing import Dict, Any


class PythonScriptExporter:
    """Generates a master Python script to mirror the original XLSX workbook."""

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        # Stores unique styles
        self.style_registry = {"borders": {}, "fills": {}, "fonts": {}, "alignments": {}}
        self.style_counters = {"BORDER": 0, "FILL": 0, "FONT": 0, "ALIGN": 0}
        self.style_map = {}  # { "prefix_hash": "ID_SEQUENCIAL" }

    def _generate_style_id(self, style_dict, prefix):
        """Generates a unique ID for a style based on its content."""
        style_str = str(sorted(style_dict.items()))
        style_hash = hashlib.md5(style_str.encode()).hexdigest()[:8]

        lookup_key = f"{prefix}_{style_hash}"
        if lookup_key in self.style_map:
            return self.style_map[lookup_key]

        self.style_counters[prefix] += 1
        seq_id = f"{prefix}_{self.style_counters[prefix]:03d}"  # Ex: FONT_001

        self.style_map[lookup_key] = seq_id

        return seq_id

    def generate_full_workbook(self, workbook_data: Dict[str, Dict[str, Any]]) -> None:
        """
        Generates a master script to build the workbook.
        Expected structure: { "SheetName": {"cells": [...], "dims": {...}, "assets": [...] } }
        """
        sheet_modules = []

        for sheet_name, content in workbook_data.items():
            clean_name = "".join(filter(str.isalnum, sheet_name))
            module_filename = f"sheet_{clean_name}.py"
            sheet_modules.append(clean_name)

            self._write_sheet_module(clean_name, sheet_name, content, module_filename)

        self._write_common_styles()
        self._write_main_reconstructor(sheet_modules)

    def _write_sheet_module(self, clean_name: str, original_name: str, content, filename: str):
        """Generate module to build the spreadsheet."""

        cmd_dims = []
        cmd_merges = []
        cmd_styles = []
        cmd_formulas = []

        # Dimensions
        dims = content.get("dims", {})
        for col, width in dims.get("cols_letter", {}).items():
            cmd_dims.append(f"    ws.column_dimensions['{col}'].width = {width}")
        for row, height in dims.get("rows_idx", {}).items():
            cmd_dims.append(f"    ws.row_dimensions[{row}].height = {height}")

        # Cells (Data & Styles)
        processed_merges = set()
        cells = content.get("cells", [])

        for c in cells:
            # Support both Dict (from JSON) and Dataclass
            data = c if isinstance(c, dict) else c.__dict__
            # Coordinate
            address = data['coordinate']            # 'A1'
            row, column = data['row'], data['col']  # (1,1)

            # Formulas
            if data.get('formula'):
                cmd_formulas.append(f"    ws['{address}'].value = '{data['formula']}'")

            # Values
            value = ""
            if data.get('value') is not None:
                value = data['value']

            # Font
            font = "None"
            font_data = {}
            if data.get('font_bold'):
                font_data['bold'] = True
            if data.get('font_italic'):
                font_data['italic'] = True
            if data.get('font_size'):
                font_data['size'] = data['font_size']
            if data.get('font_name'):
                font_data['name'] = data['font_name']

            raw_f_color = data.get('font_color')
            if isinstance(raw_f_color, str) and len(raw_f_color) <= 8:
                font_data['color'] = raw_f_color

            if font_data:
                f_id = self._generate_style_id(font_data, "FONT")
                if f_id not in self.style_registry["fonts"]:
                    f_args = [f"{k}={v}" if isinstance(v, (int, bool)) else f"{k}='{v}'"
                              for k, v in font_data.items()]
                    self.style_registry["fonts"][f_id] = f"Font({', '.join(f_args)})"
                font = f_id

            # Alignment & Rotation
            alignment = "None"
            align_data = {}
            if data.get('horizontal_align') and data.get('horizontal_align') != 'left':
                align_data['horizontal'] = data['horizontal_align']
            if data.get('vertical_align') and data.get('vertical_align') != 'bottom':
                align_data['vertical'] = data['vertical_align']
            if data.get('text_rotation', 0) != 0:
                align_data['text_rotation'] = data['text_rotation']

            if align_data:
                a_id = self._generate_style_id(align_data, "ALIGN")

                if a_id not in self.style_registry.get("alignments", {}):
                    # Se não existir a chave no dicionário principal, inicialize
                    if "alignments" not in self.style_registry:
                        self.style_registry["alignments"] = {}

                    a_args = [f"{k}={v}" if isinstance(v, int) else f"{k}='{v}'"
                              for k, v in align_data.items()]
                    self.style_registry["alignments"][a_id] = f"Alignment({', '.join(a_args)})"

                alignment = a_id

            # Borders
            border = "None"
            if data.get('borders'):
                b_id = self._generate_style_id(data['borders'], "BORDER")
                b_args = [f"{s}=Side(border_style='{d['style']}', color='{d['color']}')" for s,
                          d in data['borders'].items()]
                self.style_registry["borders"][b_id] = f"Border({', '.join(b_args)})"
                border = b_id

            # Fill
            fill = "None"
            raw_fill = data.get('fill_color')
            if isinstance(raw_fill, str) and len(raw_fill) <= 8 and " " not in raw_fill:
                fill_data = {'color': raw_fill, 'type': 'solid'}
                f_id = self._generate_style_id(fill_data, "FILL")

                if f_id not in self.style_registry["fills"]:
                    self.style_registry["fills"][f_id] = (
                        f"PatternFill(start_color='{raw_fill}', "
                        f"end_color='{raw_fill}', fill_type='solid')"
                    )
                fill = f_id

            # Style
            cmd = f"    format_cell(ws, {row}, {column}, \"\"\"{value}\"\"\", {font}, {alignment}, {border}, {fill})"
            cmd_styles.append(cmd)

            # Merge Logic
            m_range = data.get('merge_range')
            if data.get('is_merged') and m_range not in processed_merges:
                cmd_merges.append(f"    ws.merge_cells('{m_range}')")
                processed_merges.add(m_range)

        # Module
        script = [
            "# -*- coding: utf-8 -*-\n",
            "from openpyxl.styles import Alignment, Font, PatternFill, Border, Side",
            "from openpyxl.drawing.image import Image as XLImage",
            "from datetime import datetime as dt, time as tm",
            "from pathlib import Path\n",
            "# format_cell function and style constants",
            "from common import *\n\n",
            f"def build_{clean_name}(wb):",
            f"    ws = wb.create_sheet('{original_name}')"
        ]

        if cmd_dims:
            script.append("    # Dimensions")
            script.extend(cmd_dims)
        if cmd_merges:
            script.append("    # Merges")
            script.extend(cmd_merges)
        if cmd_styles:
            script.append("    # Styles")
            script.extend(cmd_styles)
        if cmd_formulas:
            script.append("    # Formulas")
            script.extend(cmd_formulas)

        # Reconstruct Assets (Images)
        assets = content.get("assets", [])
        if assets:
            script.append(f"    # Assets for {original_name}")
            for img in assets:
                script.append("    try:")
                script.append(f"        img_path = Path(__file__).parent / 'images' / '{img['filename']}'")
                script.append("        img_obj = XLImage(img_path)")
                script.append(f"        img_obj.width, img_obj.height = {img['width']}, {img['height']}")
                script.append(f"        ws.add_image(img_obj, '{img['anchor']}')")
                script.append(f"    except Exception as e: print(f' [!] Error loading image {img['filename']}: {{e}}')")

        script.append(f"    print(f\"   [+] Sheet {original_name} completed.\")\n")

        # Save Sheet module
        with open(self.output_dir / filename, "w", encoding="utf-8") as f:
            f.write("\n".join(script))

    def _write_common_styles(self):
        content = [
            "# -*- coding: utf-8 -*-\n",
            "from openpyxl.cell.cell import Cell",
            "from openpyxl.worksheet.worksheet import Worksheet",
            "from openpyxl.styles import Border, Side, PatternFill, Font, Alignment\n\n",
            "def format_cell(worksheet : Worksheet, ",
            "                row: int,",
            "                column: int,",
            "                value: str, ",
            "                font : Font = None,",
            "                alignment : Alignment = None,",
            "                border : Border = None,",
            "                fill : PatternFill = None) -> Cell:\n",
            "    cell = worksheet.cell(row=row, column=column)",
            "    if value: cell.value = value",
            "    if font: cell.font = font",
            "    if alignment: cell.alignment = alignment",
            "    if border: cell.border = border",
            "    if fill: cell.fill = fill\n",
            "    return cell\n\n",
            "# Common styles and configurations"
        ]

        for category, styles in self.style_registry.items():
            content.append(f"\n# {category.upper()}")
            for style_id in sorted(styles.keys()):
                content.append(f"{style_id} = {styles[style_id]}")

        # Save Common module
        with open(self.output_dir / "common.py", "w", encoding="utf-8") as f:
            f.write("\n".join(content))

    def _write_main_reconstructor(self, sheet_modules):
        script = [
            "# -*- coding: utf-8 -*-\n",
            "import openpyxl",
            "import sys\n",
            "from pathlib import Path\n",
            "sys.path.append(str(Path(__file__).parent))\n"
        ]

        for mod in sheet_modules:
            script.append(f"from sheet_{mod} import build_{mod}")

        script.append("\ndef main():")
        script.append("    wb = openpyxl.Workbook()")
        script.append("    if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])")

        for mod in sheet_modules:
            script.append(f"    build_{mod}(wb)")

        script.append("    output_file = 'output.xlsx'")
        script.append("    wb.save(output_file)")
        script.append("    print(f\"[!] Success: '{output_file}' generated.\")")

        script.append("\nif __name__ == '__main__':")
        script.append("    main()")

        # Save Main module
        with open(self.output_dir / "main.py", "w", encoding="utf-8") as f:
            f.write("\n".join(script))


if __name__ == "__main__":
    print("Module to convert cell metadata or JSON, generated by the Analysis Module,"
          " into a Python script using Openpyxl."
          "\nUse:\n\tfrom exporter_py import PythonScriptExporter")
